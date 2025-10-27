# routers/callcenter_router.py
from fastapi import APIRouter, Request, Form, Cookie, HTTPException
from fastapi.responses import JSONResponse
from pathlib import Path
from datetime import datetime, date, time
from typing import Optional, Dict, List, Tuple
import os
import sqlite3
from collections import Counter

import pandas as pd
from dotenv import load_dotenv

# OpenAI 為可選：沒安裝或沒金鑰也不會壞
try:
    from openai import OpenAI  # pip install openai>=1.0
except Exception:
    OpenAI = None

from app_utils import templates, get_base_context

# === Router 前綴（與原本一致） ===
router = APIRouter(prefix="/callcenter", tags=["callcenter"])

# === 載入 .env ===
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# === 舊 Excel 一次性匯入設定（維持原有功能） ===
DEFAULT_EXCEL = Path("data") / "callcenter_data.xlsx"
EXCEL_PATH = Path(os.getenv("CALLCENTER_EXCEL_PATH", str(DEFAULT_EXCEL)))
SHEET_NAME = "Tickets"

# 員工 Excel（不變）
EMP_XLSX = Path("data") / "employees.xlsx"
EMP_DEPT_COL = 4
EMP_NAME_H_COL = 8
EMP_NAME_J_COL = 10

# === SQLite 連線：優先使用 data/db.py 的 get_conn ===
try:
    from data.db import get_conn
except ImportError:
    # 後備方案：直接用 data/app.db
    DB_PATH = Path("data") / "app.db"
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    def get_conn():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn

# === 欄位、選單 ===
HEADERS = [
    "TicketID",
    "Date",
    "AnswerTime",
    "CallerName",
    "Phone",
    "Issue",
    "Dept",
    "Assignee",
    "Note",
    "Channel",
    "Status",
    "CreatedAt",
]
CHANNELS = ["電話", "LINE", "WhatsApp", "現場", "Email"]
STATUSES = ["開啟", "處理中", "待回覆", "已完成", "已轉外部"]

# === openpyxl 僅在一次性匯入時需要 ===
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ---------- employees.xlsx：動態載入部門/人員（不變） ----------
def _to_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def load_employees() -> Dict[str, List[str]]:
    depts: List[str] = []
    agents_map: Dict[str, List[str]] = {}
    if not EMP_XLSX.exists() or load_workbook is None:
        return {"depts": depts, "agentsByDept": agents_map}

    wb = load_workbook(EMP_XLSX, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = _to_text(row[EMP_DEPT_COL - 1])
        name_j = _to_text(row[EMP_NAME_J_COL - 1])
        name_h = _to_text(row[EMP_NAME_H_COL - 1])
        if not dept:
            continue
        display_name = f"{name_j} {name_h}".strip()
        if not display_name:
            continue
        agents_map.setdefault(dept, [])
        if display_name not in agents_map[dept]:
            agents_map[dept].append(display_name)
        if dept not in depts:
            depts.append(dept)

    depts.sort()
    for k, v in list(agents_map.items()):
        agents_map[k] = sorted(set(v))
    return {"depts": depts, "agentsByDept": agents_map}

# ---------- SQLite：建立資料表、索引、一次性匯入舊Excel ----------
CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS tickets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    TicketID   TEXT UNIQUE,
    Date       TEXT NOT NULL,
    AnswerTime TEXT NOT NULL,
    CallerName TEXT,
    Phone      TEXT,
    Issue      TEXT NOT NULL,
    Dept       TEXT,
    Assignee   TEXT,
    Note       TEXT,
    Channel    TEXT,
    Status     TEXT,
    CreatedAt  TEXT NOT NULL
);
"""
CREATE_INDEX_SQL = "CREATE INDEX IF NOT EXISTS idx_tickets_date ON tickets(Date);"

def ensure_db():
    with get_conn() as c:
        c.execute(CREATE_TABLE_SQL)
        c.execute(CREATE_INDEX_SQL)

def count_rows() -> int:
    with get_conn() as c:
        return c.execute("SELECT COUNT(1) AS cnt FROM tickets").fetchone()["cnt"]

def migrate_from_excel_once():
    if count_rows() > 0:
        return
    if not EXCEL_PATH.exists() or load_workbook is None:
        return
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        # 嘗試偵測表頭
        header_row = [str(v).strip() if v is not None else "" for v in rows[0]]
        start_idx = 1 if set(HEADERS).issubset(set(header_row)) else 0

        to_insert = []
        for r in rows[start_idx:]:
            r = list(r) if r else []
            d = {HEADERS[i]: (r[i] if i < len(r) else "") for i in range(len(HEADERS))}
            d["TicketID"] = str(d.get("TicketID") or "").strip() or None
            d["Date"] = _to_text(d.get("Date"))
            d["AnswerTime"] = _to_text(d.get("AnswerTime"))
            d["CallerName"] = _to_text(d.get("CallerName"))
            d["Phone"] = _to_text(d.get("Phone"))
            d["Issue"] = _to_text(d.get("Issue"))
            d["Dept"] = _to_text(d.get("Dept"))
            d["Assignee"] = _to_text(d.get("Assignee"))
            d["Note"] = _to_text(d.get("Note"))
            d["Channel"] = _to_text(d.get("Channel"))
            d["Status"] = _to_text(d.get("Status"))
            d["CreatedAt"] = _to_text(d.get("CreatedAt")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            to_insert.append(d)

        if to_insert:
            with get_conn() as c:
                c.executemany(
                    """
                    INSERT OR IGNORE INTO tickets
                    (TicketID, Date, AnswerTime, CallerName, Phone, Issue, Dept, Assignee, Note, Channel, Status, CreatedAt)
                    VALUES
                    (:TicketID, :Date, :AnswerTime, :CallerName, :Phone, :Issue, :Dept, :Assignee, :Note, :Channel, :Status, :CreatedAt)
                    """,
                    to_insert,
                )
    except Exception:
        # 匯入失敗不影響啟動
        pass

def ensure_db_and_migrate():
    ensure_db()
    migrate_from_excel_once()

# ---------- 頁面 ----------
@router.get("/")
def page(
    request: Request,
    user: Optional[str] = Cookie(default=None),
    role: Optional[str] = Cookie(default=None),
    permissions: Optional[str] = Cookie(default=None),
):
    ensure_db_and_migrate()
    ctx = get_base_context(request, user, role, permissions)
    # ➜ 新增 is_admin 旗標（role 等於 'admin' 或 permissions 內含 'admin'）
    perm_str = (permissions or "").lower()
    is_admin = (role or "").lower() == "admin" or "admin" in perm_str.split(",")
    ctx["is_admin"] = is_admin
    return templates.TemplateResponse("callcenter/callcenter.html", ctx)

# ---------- 初始化：回傳下拉所需資料 ----------
@router.get("/api/init")
def api_init():
    ensure_db_and_migrate()
    emp = load_employees()
    return {
        "depts": emp["depts"],
        "agentsByDept": emp["agentsByDept"],
        "channels": CHANNELS,
        "statuses": STATUSES,
        "headers": HEADERS,
    }

# ---------- 產生工單編號 ----------
def gen_ticket_id() -> str:
    today = datetime.now().strftime("%Y%m%d")
    prefix = f"OT-{today}-"
    with get_conn() as c:
        row = c.execute(
            "SELECT COUNT(1) AS cnt FROM tickets WHERE TicketID LIKE ?",
            (f"{prefix}%",),
        ).fetchone()
        count_today = row["cnt"] if row else 0
    return f"{prefix}{str(count_today + 1).zfill(4)}"

# ---------- 新增紀錄 ----------
@router.post("/api/ticket")
def api_ticket(
    date_: str = Form(...),
    answer_time: str = Form(...),
    caller_name: str = Form(""),
    phone: str = Form(""),
    issue: str = Form(...),
    dept: str = Form(""),
    assignee: str = Form(""),
    note: str = Form(""),
    channel: str = Form("電話"),
    status: str = Form("開啟"),
):
    ensure_db()
    try:
        _ = date.fromisoformat(date_)
    except Exception:
        return JSONResponse({"ok": False, "error": "日期格式錯誤"}, status_code=400)
    try:
        _ = time.fromisoformat(answer_time)
    except Exception:
        return JSONResponse({"ok": False, "error": "接聽時間格式錯誤"}, status_code=400)

    ticket_id = gen_ticket_id()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_conn() as c:
        c.execute(
            """
            INSERT INTO tickets
            (TicketID, Date, AnswerTime, CallerName, Phone, Issue, Dept, Assignee, Note, Channel, Status, CreatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                ticket_id,
                date_,
                answer_time,
                caller_name,
                phone,
                issue,
                dept,
                assignee,
                note,
                channel,
                status,
                created_at,
            ),
        )

    return {"ok": True, "ticketId": ticket_id}

# ---------- 列表／搜尋 ----------
@router.get("/api/list")
def api_list(limit: int = 50, q: str = ""):
    ensure_db()
    limit = max(1, min(int(limit or 50), 500))

    base_sql = "SELECT TicketID, Date, AnswerTime, CallerName, Phone, Issue, Dept, Assignee, Note, Channel, Status, CreatedAt FROM tickets"
    params: List = []

    q = (q or "").strip()
    if q:
        like = f"%{q}%"
        where = " WHERE " + " OR ".join([f"{col} LIKE ?" for col in HEADERS])
        base_sql += where
        params.extend([like] * len(HEADERS))

    base_sql += " ORDER BY id DESC LIMIT ?"
    params.append(limit)

    with get_conn() as c:
        rows = c.execute(base_sql, params).fetchall()

    dict_rows = []
    for r in rows:
        d = {HEADERS[i]: r[i] for i in range(len(HEADERS))}
        dict_rows.append(d)
    return dict_rows

# ---------- AI 客戶端（可選） ----------
def get_openai_client():
    if not OPENAI_API_KEY or not OpenAI:
        return None
    try:
        return OpenAI(api_key=OPENAI_API_KEY)
    except Exception:
        return None

def _fetch_issues_from_db() -> List[str]:
    """從 SQLite 擷取 Issue 並清理雜訊字元。"""
    ensure_db()
    with get_conn() as c:
        try:
            rows = c.execute(
                """
                SELECT TRIM(
                         REPLACE(REPLACE(REPLACE(COALESCE(Issue,''), char(9), ''),
                                         char(13), ' '),
                                         char(10), ' ')
                     ) AS IssueNorm
                FROM tickets
                WHERE COALESCE(Issue,'') <> ''
                """
            ).fetchall()
        except sqlite3.OperationalError as e:
            raise HTTPException(status_code=400, detail=f"查詢失敗：{e}")
    return [r[0] for r in rows if r and r[0]]

def _summarize_counts(issues: List[str]) -> List[Tuple[str, int]]:
    counter = Counter(issues)
    return sorted(counter.items(), key=lambda x: (-x[1], x[0]))

def _format_counts_for_prompt(items: List[Tuple[str, int]]) -> str:
    return "\n".join([f"- {k}: {v} 次" for k, v in items])

# ---------- 分析 API ----------# 先放在檔案前面某處（與其它常數一起）：
FACILITY_KEYWORDS = [
    # 停車/車位
    "車位","停車","停車費","停車場","停車位","停車折抵","停車合作","停車券",
    # 網路/充電
    "wi-fi","wifi","無線網路","網路","網速","插座","充電","電源",
    # 設施
    "健身房","健身","游泳池","泳池","spa","三溫暖","烤箱","蒸氣室","商務中心",
    "洗衣","自助洗衣","嬰兒床","加床","冰箱","吹風機","電梯","無障礙","輪椅",
    # 交通
    "接駁","接送","shuttle","叫車"
]

def _is_admin(role: Optional[str], permissions: Optional[str]) -> bool:
    r = (role or "").lower()
    pset = set([p.strip().lower() for p in (permissions or "").split(",") if p.strip()])
    return r == "admin" or "admin" in pset

def _collect_metrics():
    """回傳總筆數與服務區間（最小/最大日期字串）。"""
    with get_conn() as c:
        total = c.execute("SELECT COUNT(1) FROM tickets").fetchone()[0]
        row = c.execute(
            "SELECT MIN(Date), MAX(Date) FROM tickets WHERE COALESCE(Date,'')<>''"
        ).fetchone()
    start, end = (row[0], row[1]) if row else (None, None)
    return {"total_count": int(total or 0), "date_start": start, "date_end": end}

def _facilities_stats(issues: List[str]):
    """依關鍵字過濾『飯店設施』相關 Issue，回傳合計與 Top5。"""
    hits = []
    for s in issues:
        s_norm = (s or "").lower()
        if any(kw in s_norm for kw in FACILITY_KEYWORDS):
            hits.append(s)
    from collections import Counter
    cnt = Counter(hits)
    top = sorted(cnt.items(), key=lambda x: (-x[1], x[0]))[:5]
    return {
        "count": len(hits),
        "top": [{"issue": k, "count": v} for k, v in top]
    }

@router.post("/api/analyze")
def analyze_issues(
    use_ai: bool = True,
    role: Optional[str] = Cookie(default=None),
    permissions: Optional[str] = Cookie(default=None),
):
    # 後端權限保護
    if not _is_admin(role, permissions):
        raise HTTPException(status_code=403, detail="Forbidden: admin only")

    # 讀 DB、做統計
    issues = _fetch_issues_from_db()
    metrics = _collect_metrics()
    if not issues:
        return {
            "counts": [],
            "ai_summary": None,
            "ai_enabled": False,
            "ai_note": "資料庫沒有可分析的 Issue。",
            **metrics,
            "facilities": {"count": 0, "top": []},
        }

    counts = _summarize_counts(issues)
    payload = [{"issue": k, "count": v} for k, v in counts]

    # 設施分類
    facilities = _facilities_stats(issues)

    # （可選）AI 摘要（額度/金鑰不足會靜默）
    ai_summary = None
    ai_enabled = False
    ai_note = None
    if use_ai:
        client = get_openai_client()
        if client:
            prompt = f"""
你是一位資料分析助理。以下是客服「Issue」與次數統計，請用繁體中文產出：
1) 前三大類別與次數
2) 觀察重點（1~2行）
3) 建議的下一步（1~2行）

資料如下：
{_format_counts_for_prompt(counts)}
            """.strip()
            try:
                resp = client.chat.completions.create(
                    model="gpt-5",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.2,
                )
                ai_summary = resp.choices[0].message.content.strip()
                ai_enabled = True
            except Exception as e:
                msg = str(e).lower()
                if "insufficient_quota" in msg or "429" in msg or "quota" in msg:
                    ai_note = "AI 分析暫停：目前額度不足，已先提供統計結果。"
                else:
                    ai_note = "AI 分析暫停：設定未完成或暫時無法連線。"
        else:
            ai_note = "AI 分析暫停：尚未設定金鑰或套件未安裝。"

    return {
        "counts": payload,
        "ai_summary": ai_summary,
        "ai_enabled": ai_enabled,
        "ai_note": ai_note,
        **metrics,
        "facilities": facilities,
    }

# 健康檢查
@router.get("/api/analyze/health")
def analyze_health():
    return {
        "OPENAI_API_KEY_SET": bool(OPENAI_API_KEY),
        "OpenAIPackageInstalled": bool(OpenAI),
        "tickets_count": count_rows(),
    }
