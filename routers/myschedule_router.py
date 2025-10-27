# routers/myschedule_router.py
from __future__ import annotations
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Tuple, Optional

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app_utils import templates, get_base_context, require_permission_dep

router = APIRouter()

# ---- DB paths ----
APPROVALS_DB = Path("data/approvals.db")
PORTAL_DB = Path("data/schedule.db")


# ---- small helpers ----
@contextmanager
def _db(path: Path):
    conn = sqlite3.connect(str(path))
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


def _month_weeks(y: int, m: int) -> Tuple[date, date, List[List[date]]]:
    """回傳 (grid_start, grid_end, weeks[7])；周一開頭，補足前/後週。"""
    first = date(y, m, 1)
    start = first - timedelta(days=(first.weekday() - 0) % 7)  # 週一
    # 末日
    nxt = date(y + (m // 12), 1 if m == 12 else m + 1, 1)
    last = nxt - timedelta(days=1)
    end = last + timedelta(days=(6 - last.weekday()) % 7)
    weeks: List[List[date]] = []
    cur = start
    while cur <= end:
        row = [cur + timedelta(days=i) for i in range(7)]
        weeks.append(row)
        cur += timedelta(days=7)
    return start, end, weeks


def _month_range(y: int, m: int) -> List[date]:
    """該月份的所有日子（不跨月）。"""
    first = date(y, m, 1)
    d = first
    out = []
    while d.month == m:
        out.append(d)
        d += timedelta(days=1)
    return out


def _ensure_tables():
    """建表 + 結構演進"""
    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        # shifts
        cur.execute(
            """
        CREATE TABLE IF NOT EXISTS schedule_shifts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id TEXT NOT NULL,
            emp_name TEXT,
            department TEXT,
            department_1 TEXT,
            team TEXT,
            job_title TEXT,
            date TEXT NOT NULL,
            code TEXT,
            start_time TEXT,
            end_time TEXT,
            hours REAL,
            notes TEXT,
            created_by TEXT,
            updated_at TEXT DEFAULT (datetime('now'))
        )
        """
        )
        cols = {
            r["name"]
            for r in cur.execute("PRAGMA table_info(schedule_shifts)").fetchall()
        }
        to_add = []
        for name, typ in [
            ("department_1", "TEXT"),
            ("team", "TEXT"),
            ("job_title", "TEXT"),
            ("hours", "REAL"),
            ("notes", "TEXT"),
            ("created_by", "TEXT"),
            ("updated_at", "TEXT DEFAULT (datetime('now'))"),
        ]:
            if name not in cols:
                to_add.append((name, typ))
        for col, typ in to_add:
            try:
                cur.execute(f"ALTER TABLE schedule_shifts ADD COLUMN {col} {typ}")
            except sqlite3.OperationalError:
                pass
        try:
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_shifts_dept_date ON schedule_shifts(department, date)"
            )
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_shifts_emp_date ON schedule_shifts(emp_id, date)"
            )
        except sqlite3.OperationalError:
            pass

        # 鎖定表（改：不用表達式；用預設空字串 + 唯一索引）
        cur.execute(
            """
        CREATE TABLE IF NOT EXISTS schedule_locks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level TEXT NOT NULL,      -- 'department' | 'department_1' | 'team'
            department TEXT NOT NULL,
            department_1 TEXT DEFAULT '',   -- 改：預設空字串，避免 NULL 參與唯一性
            team TEXT DEFAULT '',           -- 改：預設空字串
            is_locked INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT DEFAULT (datetime('now'))
        )
        """
        )
        # 用唯一索引（欄位，不含表達式）來保證唯一
        cur.execute(
            """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_schedule_locks
            ON schedule_locks(level, department, department_1, team)
            """
        )
        c.commit()


_ensure_tables()


# ---------------------------
# UI: 自己部門排班
# ---------------------------
@router.get("/myschedule", response_class=JSONResponse, include_in_schema=False)
async def page_myschedule(
    request: Request,
    y: int = Query(None),
    m: int = Query(None),
    has_perm: bool = Depends(require_permission_dep("schedule_view")),
):
    today = date.today()
    y = y or today.year
    m = m or today.month
    grid_start, grid_end, weeks = _month_weeks(y, m)
    user = request.cookies.get("user")
    role = request.cookies.get("role")
    perms_raw = request.cookies.get("permissions")
    ctx = get_base_context(request, user, role, perms_raw)
    ctx.update(
        {
            "year": y,
            "month": m,
            "weeks": weeks,
            "grid_start": grid_start,
            "grid_end": grid_end,
        }
    )
    return templates.TemplateResponse("schedule/self.html", ctx)


# ---------------------------
# 三階過濾（從 approvals.db 的 Employees）
# ---------------------------
@router.get("/api/myschedule/filters", include_in_schema=False)
async def api_filters():
    with _db(APPROVALS_DB) as c:
        cur = c.cursor()
        rows = cur.execute(
            """
            SELECT DISTINCT department, department_1, team
            FROM Employees
            WHERE department IS NOT NULL AND TRIM(department) <> ''
        """
        ).fetchall()

    departments: List[str] = sorted(
        {(r["department"] or "").strip() for r in rows if r["department"]}
    )
    levels: Dict[str, List[str]] = {}
    teams: Dict[str, List[str]] = {}
    for r in rows:
        d = (r["department"] or "").strip()
        d1 = (r["department_1"] or "").strip()
        t = (r["team"] or "").strip()
        if d not in levels:
            levels[d] = []
        if d1 and d1 not in levels[d]:
            levels[d].append(d1)
        key = d + "|" + d1
        if key not in teams:
            teams[key] = []
        if t and t not in teams[key]:
            teams[key].append(t)
    for k in levels:
        levels[k].sort()
    for k in teams:
        teams[k].sort()
    return {"departments": departments, "levels": levels, "teams": teams}


# ---------------------------
# 取得員工（依部門/部級/組別）
# ---------------------------
@router.get("/api/myschedule/employees", include_in_schema=False)
async def api_employees(
    department: str,
    department_1: str = "",
    team: str = "",
):
    sql = """
    SELECT employee_id, name, english_name, job_title, department, department_1, team
    FROM Employees
    WHERE department = ?
    """
    params = [department]
    if department_1:
        sql += " AND department_1 = ?"
        params.append(department_1)
    if team:
        sql += " AND team = ?"
        params.append(team)
    sql += " ORDER BY job_title, employee_id"
    with _db(APPROVALS_DB) as c:
        cur = c.cursor()
        rs = cur.execute(sql, params).fetchall()

    out = []
    for r in rs:
        out.append(
            {
                "emp_id": r["employee_id"],
                "name": r["name"],
                "english": r["english_name"],
                "job_title": r["job_title"],
                "department": r["department"],
                "department_1": r["department_1"],
                "team": r["team"],
            }
        )
    return JSONResponse(out)


# ---------------------------
# 住房率（多來源容錯）
# ---------------------------
def _occ_band(p: float) -> str:
    if p is None:
        return ""
    if p >= 85:
        return "high"
    if p >= 60:
        return "mid"
    return "low"


@router.get("/api/myschedule/occ", include_in_schema=False)
async def api_occ(start: str, end: str):
    """
    回傳 { 'YYYY-MM-DD': {'pct': 73, 'band': 'mid'} }
    資料來源（依序嘗試）：
      1) schedule_occupancy(date TEXT PRIMARY KEY, pct REAL) in schedule.db
      2) routers.calendar_router 中的函式（若專案有提供）
      3) 找不到 → 回傳空值
    """
    dt0 = date.fromisoformat(start)
    dt1 = date.fromisoformat(end)
    days = []
    d = dt0
    while d <= dt1:
        days.append(d.isoformat())
        d += timedelta(days=1)

    result: Dict[str, Dict[str, Optional[float]]] = {d: None for d in days}

    # 1) schedule_occupancy
    try:
        with _db(PORTAL_DB) as c:
            cur = c.cursor()
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS schedule_occupancy(
                    date TEXT PRIMARY KEY,
                    pct REAL
                )
            """
            )
            rows = cur.execute(
                "SELECT date, pct FROM schedule_occupancy WHERE date BETWEEN ? AND ?",
                (start, end),
            ).fetchall()
            for r in rows:
                pct = float(r["pct"] or 0.0)
                result[r["date"]] = {"pct": round(pct), "band": _occ_band(pct)}
    except Exception:
        pass

    # 2) 嘗試從 calendar_router 取得
    try:
        if any(result[d] is None for d in days):
            # 允許你的 calendar_router 提供 get_occupancy_map(start,end) 或 calc_occupancy_range
            from routers import calendar_router as cal

            occ_map = None
            if hasattr(cal, "get_occupancy_map"):
                occ_map = cal.get_occupancy_map(start, end)  # type: ignore[attr-defined]
            elif hasattr(cal, "calc_occupancy_range"):
                occ_map = cal.calc_occupancy_range(start, end)  # type: ignore[attr-defined]
            if isinstance(occ_map, dict):
                for k, v in occ_map.items():
                    if k in result and v is not None:
                        pct = float(v)
                        result[k] = {"pct": round(pct), "band": _occ_band(pct)}
    except Exception:
        pass

    # 3) 補空
    out = {}
    for d in days:
        if isinstance(result.get(d), dict):
            out[d] = result[d]
        else:
            out[d] = None
    return JSONResponse(out)


# ---------------------------
# 讀取班表
# ---------------------------
@router.get("/api/myschedule/shifts", include_in_schema=False)
async def api_shifts(
    start: str,
    end: str,
    department: str,
    department_1: str = "",
    team: str = "",
):
    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        sql = """SELECT id, emp_id, emp_name, department, department_1, team, job_title, date, code,
                        start_time, end_time, hours
                 FROM schedule_shifts
                 WHERE department=? AND date BETWEEN date(?) AND date(?)"""
        params = [department, start, end]
        if department_1:
            sql += " AND department_1=?"
            params.append(department_1)
        if team:
            sql += " AND team=?"
            params.append(team)
        sql += " ORDER BY emp_id, date"
        rows = cur.execute(sql, params).fetchall()
    out = [dict(r) for r in rows]
    return JSONResponse(out)


# ---------------------------
# 鎖定機制（查詢 + 設定 + 內部判斷）
# ---------------------------
def _lock_effective(
    department: str, department_1: str = "", team: str = ""
) -> Tuple[bool, Optional[str]]:
    """回傳 (locked, level)；team > department_1 > department"""
    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        if team:
            r = cur.execute(
                """SELECT is_locked FROM schedule_locks
                               WHERE level='team' AND department=? AND (department_1=? OR department_1 IS NULL OR department_1='')
                                     AND team=?""",
                (department, department_1, team),
            ).fetchone()
            if r and (r["is_locked"] == 1):
                return True, "team"
        if department_1:
            r = cur.execute(
                """SELECT is_locked FROM schedule_locks
                               WHERE level='department_1' AND department=? AND department_1=?""",
                (department, department_1),
            ).fetchone()
            if r and (r["is_locked"] == 1):
                return True, "department_1"
        r = cur.execute(
            """SELECT is_locked FROM schedule_locks
                           WHERE level='department' AND department=?""",
            (department,),
        ).fetchone()
        if r and (r["is_locked"] == 1):
            return True, "department"
    return False, None


@router.get("/api/myschedule/lock", include_in_schema=False)
async def api_lock_status(department: str, department_1: str = "", team: str = ""):
    locked, level = _lock_effective(department, department_1, team)
    return {"locked": locked, "level": level}


@router.post("/api/myschedule/lock", include_in_schema=False)
async def api_lock_set(
    level: str = Form(...),  # 'department' | 'department_1' | 'team'
    department: str = Form(...),
    department_1: str = Form(""),
    team: str = Form(""),
    is_locked: int = Form(...),  # 1=鎖定, 0=開啟
):
    if level not in ("department", "department_1", "team"):
        raise HTTPException(400, "invalid level")
    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        if level == "team":
            cur.execute(
                """UPDATE schedule_locks SET is_locked=?, updated_at=datetime('now')
                           WHERE level='team' AND department=? AND (department_1=? OR department_1 IS NULL OR department_1='')
                                 AND team=?""",
                (int(is_locked), department, department_1, team),
            )
            if cur.rowcount == 0:
                cur.execute(
                    """INSERT INTO schedule_locks(level,department,department_1,team,is_locked)
                               VALUES('team',?,?,?,?)""",
                    (department, department_1, team, int(is_locked)),
                )
        elif level == "department_1":
            cur.execute(
                """UPDATE schedule_locks SET is_locked=?, updated_at=datetime('now')
                           WHERE level='department_1' AND department=? AND department_1=?""",
                (int(is_locked), department, department_1),
            )
            if cur.rowcount == 0:
                cur.execute(
                    """INSERT INTO schedule_locks(level,department,department_1,is_locked)
                               VALUES('department_1',?,?,?)""",
                    (department, department_1, int(is_locked)),
                )
        else:
            cur.execute(
                """UPDATE schedule_locks SET is_locked=?, updated_at=datetime('now')
                           WHERE level='department' AND department=?""",
                (int(is_locked), department),
            )
            if cur.rowcount == 0:
                cur.execute(
                    """INSERT INTO schedule_locks(level,department,is_locked)
                               VALUES('department',?,?)""",
                    (department, int(is_locked)),
                )
        c.commit()
    return {"ok": True}


# ---------------------------
# 新增 / 更新 / 刪除 班表（含鎖定檢查）
# ---------------------------
@router.post("/api/myschedule/shifts", include_in_schema=False)
async def api_shift_create(
    emp_id: str = Form(...),
    emp_name: str = Form(""),
    job_title: str = Form(""),
    department: str = Form(...),
    department_1: str = Form(""),
    team: str = Form(""),
    date_s: str = Form(...),
    code: str = Form(""),
    start_time: str = Form(""),
    end_time: str = Form(""),
    hours: float = Form(None),
    created_by: str = Form(""),
):
    # 鎖定時禁止維護
    locked, lvl = _lock_effective(department, department_1, team)
    if locked:
        raise HTTPException(status_code=423, detail=f"locked by {lvl}")

    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        cur.execute(
            """INSERT INTO schedule_shifts
            (emp_id, emp_name, job_title, department, department_1, team, date, code, start_time, end_time, hours, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                emp_id,
                emp_name,
                job_title,
                department,
                department_1,
                team,
                date_s,
                code,
                start_time,
                end_time,
                hours,
                created_by,
            ),
        )
        c.commit()
        return JSONResponse({"ok": True, "id": cur.lastrowid})


@router.put("/api/myschedule/shifts/{sid}", include_in_schema=False)
async def api_shift_update(
    sid: int,
    code: str = Form(...),
    start_time: str = Form(""),
    end_time: str = Form(""),
    hours: float = Form(None),
):
    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        row = cur.execute(
            "SELECT department, department_1, team FROM schedule_shifts WHERE id=?",
            (sid,),
        ).fetchone()
        if not row:
            raise HTTPException(404, "shift not found")
        locked, lvl = _lock_effective(
            row["department"], row["department_1"] or "", row["team"] or ""
        )
        if locked:
            raise HTTPException(status_code=423, detail=f"locked by {lvl}")

        cur.execute(
            """UPDATE schedule_shifts
                       SET code=?, start_time=?, end_time=?, hours=?, updated_at=datetime('now')
                       WHERE id=?""",
            (code, start_time, end_time, hours, sid),
        )
        c.commit()
        return JSONResponse({"ok": True, "id": sid})


@router.delete("/api/myschedule/shifts/{sid}", include_in_schema=False)
async def api_shift_delete(sid: int):
    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        row = cur.execute(
            "SELECT department, department_1, team FROM schedule_shifts WHERE id=?",
            (sid,),
        ).fetchone()
        if not row:
            return JSONResponse({"ok": True, "id": sid})
        locked, lvl = _lock_effective(
            row["department"], row["department_1"] or "", row["team"] or ""
        )
        if locked:
            raise HTTPException(status_code=423, detail=f"locked by {lvl}")
        cur.execute("DELETE FROM schedule_shifts WHERE id=?", (sid,))
        c.commit()
        return JSONResponse({"ok": True, "id": sid})


# ---------------------------
# 匯出 Excel（UTF-8 檔名安全）
# ---------------------------
import urllib.parse


@router.get("/myschedule/export", include_in_schema=False)
async def myschedule_export(
    y: int,
    m: int,
    department: str,
    department_1: str = "",
    team: str = "",
):
    # 1) 員工
    with _db(APPROVALS_DB) as c:
        cur = c.cursor()
        sql = """SELECT employee_id, name, english_name, job_title, department, department_1, team
                 FROM Employees WHERE department=?"""
        params = [department]
        if department_1:
            sql += " AND department_1=?"
            params.append(department_1)
        if team:
            sql += " AND team=?"
            params.append(team)
        sql += " ORDER BY job_title, employee_id"
        emps = cur.execute(sql, params).fetchall()

    # 2) 班表
    days = _month_range(y, m)
    start_s, end_s = days[0].isoformat(), days[-1].isoformat()
    with _db(PORTAL_DB) as c:
        cur = c.cursor()
        sql = """SELECT emp_id, date, code FROM schedule_shifts
                 WHERE department=? AND date BETWEEN date(?) AND date(?)"""
        params = [department, start_s, end_s]
        if department_1:
            sql += " AND department_1=?"
            params.append(department_1)
        if team:
            sql += " AND team=?"
            params.append(team)
        rows = cur.execute(sql, params).fetchall()
    shifts = {(r["emp_id"], r["date"]): (r["code"] or "") for r in rows}

    # 3) 產生 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}-{m:02d}"

    headers = ["職稱", "員編", "姓名"]
    for d in days:
        headers.append(f"{d.month}/{d.day}")
    headers.append("合計(天/時)")
    ws.append(headers)

    def hours_of(code: str) -> float:
        # 你的班別工時規則（同前端預設）
        return 8.0 if code and code not in ("休", "例") else 0.0

    for e in emps:
        emp_id = e["employee_id"]
        name = f'{e["english_name"] or ""} {e["name"] or ""}'.strip()
        row = [e["job_title"] or "", emp_id, name]
        work_days = rest_days = hol_days = 0
        hours = 0.0
        for d in days:
            code = shifts.get((emp_id, d.isoformat()), "")
            row.append(code)
            if code == "休":
                rest_days += 1
            elif code == "例":
                hol_days += 1
            elif code:
                work_days += 1
                hours += hours_of(code)
        row.append(f"上班{work_days}/工時{int(hours)}")
        ws.append(row)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions[get_column_letter(len(headers))].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    display_name = f"排班_{department}_{y}-{m:02d}.xlsx"
    ascii_name = f"schedule_{y}{m:02d}.xlsx"
    encoded_name = urllib.parse.quote(display_name)
    cd = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )
