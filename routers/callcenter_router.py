from fastapi import APIRouter, Request, Form, Cookie
from fastapi.responses import JSONResponse
from pathlib import Path
from datetime import datetime, date, time
from typing import Optional, Dict, List
import os
import sqlite3

# 只在「一次性資料轉移」時會用到 openpyxl
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # 若環境沒有也不致命（僅影響自動匯入舊Excel）

from app_utils import templates, get_base_context

# 與原本一致的 Router 前綴
router = APIRouter(prefix="/callcenter", tags=["callcenter"])

# === 舊 Excel 資料路徑（用於一次性自動匯入） ===
DEFAULT_EXCEL = Path("data") / "callcenter_data.xlsx"
EXCEL_PATH = Path(os.getenv("CALLCENTER_EXCEL_PATH", str(DEFAULT_EXCEL)))
SHEET_NAME = "Tickets"

# === 員工 Excel（動態來源，不變） ===
EMP_XLSX = Path("data") / "employees.xlsx"
EMP_DEPT_COL = 4   # D 欄 = 部門
EMP_NAME_H_COL = 8 # H 欄 = 名稱片段
EMP_NAME_J_COL = 10

# === SQLite 連線（沿用專案的 data.db 模組習慣） ===
# 若你專案已有 data/db.py 的 get_conn()，可直接 import 使用：
try:
    from data.db import get_conn  # 與 auth_router 相同使用方式
except ImportError:
    # 後備方案：直接對 data/auth.db 或 data/app.db 建立連線
    DB_PATH = Path("data") / "app.db"
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    def get_conn():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn

# 工單欄位（維持與前端/舊檔一致）
HEADERS = [
    "TicketID", "Date", "AnswerTime", "CallerName", "Phone", "Issue",
    "Dept", "Assignee", "Note", "Channel", "Status", "CreatedAt",
]

# 其它下拉選單（不變）
CHANNELS = ["電話", "LINE", "WhatsApp", "現場", "Email"]
STATUSES = ["開啟", "處理中", "待回覆", "已完成", "已轉外部"]

# ---------- employees.xlsx：動態載入部門/人員（不變） ----------
def _to_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def load_employees() -> Dict[str, List[str]]:
    depts: List[str] = []
    agents_map: Dict[str, List[str]] = {}

    if not EMP_XLSX.exists():
        return {"depts": depts, "agentsByDept": agents_map}

    if load_workbook is None:
        return {"depts": depts, "agentsByDept": agents_map}

    wb = load_workbook(EMP_XLSX, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        dept   = _to_text(row[EMP_DEPT_COL - 1])
        name_j = _to_text(row[EMP_NAME_J_COL - 1])
        name_h = _to_text(row[EMP_NAME_H_COL - 1])
        if not dept:
            continue

        display_name = f"{name_j} {name_h}".strip()
        if not display_name:
            continue

        agents_map.setdefault(dept, [])
        if display_name not in agents_map[dept]:
            agents_map[dept].append(display_name)

        if dept not in depts:
            depts.append(dept)

    depts.sort()
    for k, v in list(agents_map.items()):
        agents_map[k] = sorted(set(v))
    return {"depts": depts, "agentsByDept": agents_map}

# ---------- SQLite：建立資料表、索引、一次性匯入舊Excel ----------
CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS tickets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    TicketID   TEXT UNIQUE,
    Date       TEXT NOT NULL,
    AnswerTime TEXT NOT NULL,
    CallerName TEXT,
    Phone      TEXT,
    Issue      TEXT NOT NULL,
    Dept       TEXT,
    Assignee   TEXT,
    Note       TEXT,
    Channel    TEXT,
    Status     TEXT,
    CreatedAt  TEXT NOT NULL
);
"""
CREATE_INDEX_SQL = """
CREATE INDEX IF NOT EXISTS idx_tickets_date ON tickets(Date);
"""

def ensure_db():
    with get_conn() as c:
        c.execute(CREATE_TABLE_SQL)
        c.execute(CREATE_INDEX_SQL)

def count_rows() -> int:
    with get_conn() as c:
        return c.execute("SELECT COUNT(1) AS cnt FROM tickets").fetchone()["cnt"]

def migrate_from_excel_once():
    """
    若 tickets 為空且存在舊 Excel，將其匯入 SQLite。
    僅執行一次；若沒有 openpyxl 或檔案不存在則略過。
    """
    if count_rows() > 0:
        return
    if not EXCEL_PATH.exists() or load_workbook is None:
        return

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        # 讀取第一列表頭，嘗試對應到 HEADERS
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        # 嘗試偵測表頭列（容錯：若第一列不是表頭，當作沒有表頭直接照順序對應）
        header_row = [str(v).strip() if v is not None else "" for v in rows[0]]
        start_idx = 1
        if set(HEADERS).issubset(set(header_row)):
            start_idx = 1
        else:
            # 沒偵測到完整表頭，當作資料從第1列開始
            start_idx = 0

        to_insert = []
        for r in rows[start_idx:]:
            r = list(r) if r else []
            d = {HEADERS[i]: (r[i] if i < len(r) else "") for i in range(len(HEADERS))}
            # 基本清理
            d["TicketID"] = str(d.get("TicketID") or "").strip() or None
            d["Date"] = _to_text(d.get("Date"))
            d["AnswerTime"] = _to_text(d.get("AnswerTime"))
            d["CallerName"] = _to_text(d.get("CallerName"))
            d["Phone"] = _to_text(d.get("Phone"))
            d["Issue"] = _to_text(d.get("Issue"))
            d["Dept"] = _to_text(d.get("Dept"))
            d["Assignee"] = _to_text(d.get("Assignee"))
            d["Note"] = _to_text(d.get("Note"))
            d["Channel"] = _to_text(d.get("Channel"))
            d["Status"] = _to_text(d.get("Status"))
            d["CreatedAt"] = _to_text(d.get("CreatedAt")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            to_insert.append(d)

        if to_insert:
            with get_conn() as c:
                c.executemany("""
                    INSERT OR IGNORE INTO tickets
                    (TicketID, Date, AnswerTime, CallerName, Phone, Issue, Dept, Assignee, Note, Channel, Status, CreatedAt)
                    VALUES
                    (:TicketID, :Date, :AnswerTime, :CallerName, :Phone, :Issue, :Dept, :Assignee, :Note, :Channel, :Status, :CreatedAt)
                """, to_insert)
    except Exception:
        # 匯入失敗不影響啟動；可在日後以工具手動匯入
        pass

def ensure_db_and_migrate():
    ensure_db()
    migrate_from_excel_once()

# ---------- 頁面 ----------
@router.get("/")
def page(
    request: Request,
    user: Optional[str] = Cookie(default=None),
    role: Optional[str] = Cookie(default=None),
    permissions: Optional[str] = Cookie(default=None),
):
    ensure_db_and_migrate()
    ctx = get_base_context(request, user, role, permissions)
    # 與前端模板路徑一致（保持不變）
    return templates.TemplateResponse("callcenter/callcenter.html", ctx)

# ---------- 初始化：回傳下拉所需資料 ----------
@router.get("/api/init")
def api_init():
    ensure_db_and_migrate()
    emp = load_employees()
    return {
        "depts": emp["depts"],
        "agentsByDept": emp["agentsByDept"],
        "channels": CHANNELS,
        "statuses": STATUSES,
        "headers": HEADERS,
    }

# ---------- 產生工單編號（維持 OT-YYYYMMDD-#### 格式） ----------
def gen_ticket_id() -> str:
    today = datetime.now().strftime("%Y%m%d")
    prefix = f"OT-{today}-"
    with get_conn() as c:
        row = c.execute(
            "SELECT COUNT(1) AS cnt FROM tickets WHERE TicketID LIKE ?",
            (f"{prefix}%",)
        ).fetchone()
        count_today = row["cnt"] if row else 0
    return f"{prefix}{str(count_today + 1).zfill(4)}"

# ---------- 新增紀錄 ----------
@router.post("/api/ticket")
def api_ticket(
    date_: str = Form(...),
    answer_time: str = Form(...),
    caller_name: str = Form(""),
    phone: str = Form(""),
    issue: str = Form(...),
    dept: str = Form(""),
    assignee: str = Form(""),
    note: str = Form(""),
    channel: str = Form("電話"),
    status: str = Form("開啟"),
):
    ensure_db()

    # 後端驗證：日期/時間格式
    try:
        _ = date.fromisoformat(date_)
    except Exception:
        return JSONResponse({"ok": False, "error": "日期格式錯誤"}, status_code=400)
    try:
        _ = time.fromisoformat(answer_time)
    except Exception:
        return JSONResponse({"ok": False, "error": "接聽時間格式錯誤"}, status_code=400)

    ticket_id = gen_ticket_id()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_conn() as c:
        c.execute("""
            INSERT INTO tickets
            (TicketID, Date, AnswerTime, CallerName, Phone, Issue, Dept, Assignee, Note, Channel, Status, CreatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            ticket_id, date_, answer_time, caller_name, phone, issue,
            dept, assignee, note, channel, status, created_at,
        ))

    return {"ok": True, "ticketId": ticket_id}

# ---------- 列表／搜尋（維持回傳格式與欄位順序） ----------
@router.get("/api/list")
def api_list(limit: int = 50, q: str = ""):
    ensure_db()
    limit = max(1, min(int(limit or 50), 500))

    base_sql = "SELECT TicketID, Date, AnswerTime, CallerName, Phone, Issue, Dept, Assignee, Note, Channel, Status, CreatedAt FROM tickets"
    params: List = []

    # 關鍵字跨欄位 LIKE
    q = (q or "").strip()
    if q:
        like = f"%{q}%"
        where = " WHERE " + " OR ".join([f"{col} LIKE ?" for col in HEADERS])
        base_sql += where
        params.extend([like] * len(HEADERS))

    # 以 id DESC 讓最新在前
    base_sql += " ORDER BY id DESC LIMIT ?"
    params.append(limit)

    with get_conn() as c:
        rows = c.execute(base_sql, params).fetchall()

    # 轉回指定欄位的 dict 列表（與前端渲染一致）
    dict_rows = []
    for r in rows:
        d = {HEADERS[i]: r[i] for i in range(len(HEADERS))}
        dict_rows.append(d)

    return dict_rows
