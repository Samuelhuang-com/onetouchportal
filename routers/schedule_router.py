# routers/schedule_router.py
from fastapi import APIRouter, Request, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse
from pathlib import Path
from sqlite3 import connect, OperationalError
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple
import json, math, re

from app_utils import templates, get_base_context
from data.perm_service import has_permission

router = APIRouter()

# === DB paths ===
PORTAL_DB = Path("data/schedule.db")  # ← 你指定的新 DB
APPROVALS_DB = Path("data/approvals.db")
MSR02_TABLE = "MSR02"


def _db(path: Path):
    conn = connect(path)
    conn.row_factory = lambda cur, row: {
        cur.description[i][0]: row[i] for i in range(len(row))
    }
    return conn


# === permissions ===
def _cookie_perms(request: Request) -> dict:
    try:
        return json.loads(request.cookies.get("permissions") or "{}")
    except Exception:
        return {}


def _has_perm(request: Request, key: str) -> bool:
    perms = _cookie_perms(request)
    if perms.get(key):
        return True
    user = request.cookies.get("user") or ""
    return has_permission(user, key)


def _require_perm(request: Request, key: str):
    if not _has_perm(request, key):
        raise HTTPException(status_code=401, detail=f"permission required: {key}")


# === helpers ===
_num_re = re.compile(r"[-+]?\d*\.?\d+")


def _num(val: Optional[str]) -> Optional[float]:
    if val is None:
        return None
    s = str(val).replace(",", "").strip()
    m = _num_re.search(s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except:
        return None


def _ratio(val: Optional[str]) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        if s.endswith("%"):
            return float(s[:-1]) / 100.0
        v = float(s)
        if 0.0 <= v <= 1.0:
            return v
        if 1.0 < v <= 100.0:
            return v / 100.0
        return None
    except:
        return None


def _band(r: Optional[float]) -> Optional[str]:
    if r is None:
        return None
    if r < 0.60:
        return "low"
    if r < 0.80:
        return "mid"
    return "high"


def _ensure_tables():
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        c.execute(
            """
        CREATE TABLE IF NOT EXISTS schedule_shifts(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id TEXT NOT NULL,
            emp_name TEXT,
            department TEXT,
            department_1 TEXT,
            role TEXT,
            date TEXT NOT NULL,
            code TEXT,
            start_time TEXT,
            end_time TEXT,
            location TEXT,
            notes TEXT,
            created_by TEXT,
            updated_at TEXT DEFAULT (datetime('now'))
        )"""
        )
        c.execute(
            """
        CREATE TABLE IF NOT EXISTS schedule_rules(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department TEXT NOT NULL,
            role TEXT NOT NULL,
            base REAL DEFAULT 0,
            per_occ REAL DEFAULT 0,
            per_guest REAL DEFAULT 0,
            per_breakfast REAL DEFAULT 0,
            min_staff INTEGER DEFAULT 0,
            max_staff INTEGER DEFAULT 9999,
            note TEXT
        )"""
        )
        c.execute(
            """
        CREATE TABLE IF NOT EXISTS schedule_day_metrics(
            date TEXT PRIMARY KEY,
            occ_ratio REAL,
            guest_count INTEGER,
            breakfast_count INTEGER
        )"""
        )
        conn.commit()


_ensure_tables()


# === Occupancy loader（沿用 Calendar 的邏輯） ===
def _load_month_occupancy(
    grid_start: date, grid_end: date
) -> Dict[date, Dict[str, Any]]:
    result: Dict[date, Dict[str, Any]] = {}
    if not PORTAL_DB.exists():
        return result
    start_key = grid_start.strftime("%Y%m%d")
    end_key = grid_end.strftime("%Y%m%d")
    try:
        with _db(PORTAL_DB) as conn:
            cur = conn.cursor()
            rows = cur.execute(
                f'SELECT * FROM "{MSR02_TABLE}" WHERE "Date" BETWEEN ? AND ? ORDER BY "Date"',
                (start_key, end_key),
            ).fetchall()
            for r in rows:
                ymd = str(r.get("Date") or "").strip()
                if len(ymd) != 8:
                    continue
                try:
                    d = date(int(ymd[0:4]), int(ymd[4:6]), int(ymd[6:8]))
                except:
                    continue

                ratio = (
                    _ratio(r.get("occ_total"))
                    or _ratio(r.get("occ_sellable_rooms"))
                    or _ratio(r.get("occ_total_rooms"))
                )

                if ratio is None:
                    num = _num(r.get("rooms_occupied_total")) or _num(
                        r.get("rooms_sold")
                    )
                    den = _num(r.get("sellable_rooms"))
                    if den is None:
                        total = _num(r.get("total_rooms"))
                        withdrawn = sum(
                            _num(r.get(k)) or 0.0
                            for k in (
                                "ooo_rooms",
                                "mbk_rooms",
                                "hus_rooms",
                                "comp_rooms",
                            )
                        )
                        if total is not None:
                            den = max(total - withdrawn, 0.0)
                    if num is not None and den and den > 0:
                        ratio = float(num) / float(den)

                if ratio is None:
                    continue
                result[d] = {
                    "ratio": ratio,
                    "band": _band(ratio),
                    "display": f"{round(ratio*100,1)}%",
                }
        return result
    except OperationalError:
        return result


# === Employees & 部門清單 ===
def _load_employees(
    dept: Optional[str] = None, dept1: Optional[str] = None
) -> List[Dict[str, Any]]:
    table_names = ["Employees", "employees"]
    rows: List[Dict[str, Any]] = []
    if not APPROVALS_DB.exists():
        return rows
    with _db(APPROVALS_DB) as conn:
        c = conn.cursor()
        for t in table_names:
            try:
                sql = f"""
                SELECT
                  COALESCE(emp_no, id) as emp_id,
                  COALESCE(name, chinese_name, display_name) as name,
                  COALESCE(english_name, en_name) as english_name,
                  department, department_1,
                  COALESCE(title, job_title) as title,
                  email
                FROM {t}
                WHERE (status IS NULL OR status NOT IN ('left','離職','disabled'))
                """
                params: List[Any] = []
                if dept:
                    sql += " AND department = ?"
                    params.append(dept)
                if dept1:
                    sql += " AND department_1 = ?"
                    params.append(dept1)
                part = c.execute(sql, tuple(params)).fetchall()
                rows.extend(part)
                break
            except OperationalError:
                continue
    return rows


def _load_dept_lists() -> Tuple[List[str], List[str]]:
    depts, dept1s = [], []
    if not APPROVALS_DB.exists():
        return depts, dept1s
    with _db(APPROVALS_DB) as conn:
        c = conn.cursor()
        for t in ("Employees", "employees"):
            try:
                d = [
                    r["department"]
                    for r in c.execute(
                        f"SELECT DISTINCT department FROM {t} WHERE department IS NOT NULL"
                    ).fetchall()
                    if r["department"]
                ]
                d1 = [
                    r["department_1"]
                    for r in c.execute(
                        f"SELECT DISTINCT department_1 FROM {t} WHERE department_1 IS NOT NULL"
                    ).fetchall()
                    if r["department_1"]
                ]
                depts = sorted(set(d))
                dept1s = sorted(set(d1))
                break
            except OperationalError:
                continue
    return depts, dept1s


@router.get("/api/schedule/dept-options", include_in_schema=False)
async def api_dept_options(request: Request):
    _require_perm(request, "schedule_view")
    d, d1 = _load_dept_lists()
    return JSONResponse({"departments": d, "department_1": d1})


@router.get("/api/schedule/employees", include_in_schema=False)
async def api_employees(
    request: Request,
    department: Optional[str] = None,
    department_1: Optional[str] = None,
):
    _require_perm(request, "schedule_view")
    return JSONResponse(_load_employees(department, department_1))


# === Shifts CRUD ===
@router.get("/api/schedule/shifts", include_in_schema=False)
async def api_shifts_list(
    request: Request,
    start: str,
    end: str,
    department: Optional[str] = None,
    department_1: Optional[str] = None,
):
    _require_perm(request, "schedule_view")
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        sql = "SELECT * FROM schedule_shifts WHERE date BETWEEN date(?) AND date(?)"
        params = [start, end]
        if department:
            sql += " AND department = ?"
            params.append(department)
        if department_1:
            sql += " AND department_1 = ?"
            params.append(department_1)
        sql += " ORDER BY date ASC, department, department_1, emp_name"
        rows = c.execute(sql, tuple(params)).fetchall()
    return JSONResponse(rows)


@router.post("/api/schedule/shifts", include_in_schema=False)
async def api_shifts_create(
    request: Request,
    emp_id: str = Form(...),
    emp_name: str = Form(...),
    date_s: str = Form(...),
    code: str = Form(""),
    start_time: str = Form(None),
    end_time: str = Form(None),
    department: str = Form(None),
    department_1: str = Form(None),
    role: str = Form(None),
    location: str = Form(None),
    notes: str = Form(None),
):
    _require_perm(request, "schedule_manage")
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        c.execute(
            """
        INSERT INTO schedule_shifts (emp_id, emp_name, department, department_1, role, date, code, start_time, end_time, location, notes, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                emp_id,
                emp_name,
                department,
                department_1,
                role,
                date_s,
                code or None,
                start_time,
                end_time,
                location,
                notes,
                request.cookies.get("user") or "system",
            ),
        )
        conn.commit()
        new_id = c.lastrowid
        row = c.execute(
            "SELECT * FROM schedule_shifts WHERE id=?", (new_id,)
        ).fetchone()
    return JSONResponse(row)


@router.put("/api/schedule/shifts/{sid}", include_in_schema=False)
async def api_shifts_update(
    request: Request,
    sid: int,
    code: str = Form(None),
    start_time: str = Form(None),
    end_time: str = Form(None),
    role: str = Form(None),
    location: str = Form(None),
    notes: str = Form(None),
):
    _require_perm(request, "schedule_manage")
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        c.execute(
            """
        UPDATE schedule_shifts
        SET code = COALESCE(?, code),
            start_time = COALESCE(?, start_time),
            end_time = COALESCE(?, end_time),
            role = COALESCE(?, role),
            location = COALESCE(?, location),
            notes = COALESCE(?, notes),
            updated_at = datetime('now')
        WHERE id = ?
        """,
            (code, start_time, end_time, role, location, notes, sid),
        )
        conn.commit()
        row = c.execute("SELECT * FROM schedule_shifts WHERE id=?", (sid,)).fetchone()
    return JSONResponse(row)


@router.delete("/api/schedule/shifts/{sid}", include_in_schema=False)
async def api_shifts_delete(request: Request, sid: int):
    _require_perm(request, "schedule_manage")
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        c.execute("DELETE FROM schedule_shifts WHERE id=?", (sid,))
        conn.commit()
    return JSONResponse({"ok": True, "id": sid})


# === Rules CRUD ===
@router.get("/api/schedule/rules", include_in_schema=False)
async def api_rules_list(request: Request, department: Optional[str] = None):
    _require_perm(request, "schedule_view")
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        if department:
            rows = c.execute(
                "SELECT * FROM schedule_rules WHERE department=? ORDER BY role",
                (department,),
            ).fetchall()
        else:
            rows = c.execute(
                "SELECT * FROM schedule_rules ORDER BY department, role"
            ).fetchall()
    return JSONResponse(rows)


@router.post("/api/schedule/rules", include_in_schema=False)
async def api_rules_upsert(
    request: Request,
    department: str = Form(...),
    role: str = Form(...),
    base: float = Form(0),
    per_occ: float = Form(0),
    per_guest: float = Form(0),
    per_breakfast: float = Form(0),
    min_staff: int = Form(0),
    max_staff: int = Form(9999),
    note: str = Form(None),
):
    _require_perm(request, "schedule_manage")
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        exist = c.execute(
            "SELECT id FROM schedule_rules WHERE department=? AND role=?",
            (department, role),
        ).fetchone()
        if exist:
            c.execute(
                """
            UPDATE schedule_rules SET base=?, per_occ=?, per_guest=?, per_breakfast=?, min_staff=?, max_staff=?, note=?
            WHERE id=?
            """,
                (
                    base,
                    per_occ,
                    per_guest,
                    per_breakfast,
                    min_staff,
                    max_staff,
                    note,
                    exist["id"],
                ),
            )
            rid = exist["id"]
        else:
            c.execute(
                """
            INSERT INTO schedule_rules (department, role, base, per_occ, per_guest, per_breakfast, min_staff, max_staff, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                (
                    department,
                    role,
                    base,
                    per_occ,
                    per_guest,
                    per_breakfast,
                    min_staff,
                    max_staff,
                    note,
                ),
            )
            rid = c.lastrowid
        conn.commit()
        row = c.execute("SELECT * FROM schedule_rules WHERE id=?", (rid,)).fetchone()
    return JSONResponse(row)


# === Day metrics ===
@router.get("/api/schedule/day-metrics", include_in_schema=False)
async def api_day_metrics(request: Request, start: str, end: str):
    _require_perm(request, "schedule_view")
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        rows = c.execute(
            """
        SELECT * FROM schedule_day_metrics
        WHERE date BETWEEN date(?) AND date(?)
        ORDER BY date ASC
        """,
            (start, end),
        ).fetchall()
    return JSONResponse(rows)


# === Occupancy API（供前端拉） ===
@router.get("/api/schedule/occ", include_in_schema=False)
async def api_schedule_occ(request: Request, start: str, end: str):
    _require_perm(request, "schedule_view")
    start_d = datetime.fromisoformat(start).date()
    end_d = datetime.fromisoformat(end).date()
    raw = _load_month_occupancy(start_d, end_d)
    out = {}
    cur = start_d
    while cur <= end_d:
        dmeta = raw.get(cur)
        if dmeta and dmeta.get("ratio") is not None:
            r = float(dmeta["ratio"])
            out[cur.isoformat()] = {
                "ratio": r,
                "pct": round(r * 100, 1),
                "band": dmeta.get("band"),
            }
        cur += timedelta(days=1)
    return JSONResponse(out)


# === 人力評估 ===
@router.get("/api/schedule/assess", include_in_schema=False)
async def api_schedule_assess(
    request: Request, start: str, end: str, department: Optional[str] = None
):
    _require_perm(request, "schedule_view")
    start_d = datetime.fromisoformat(start).date()
    end_d = datetime.fromisoformat(end).date()
    occ_map = _load_month_occupancy(start_d, end_d)

    with _db(PORTAL_DB) as conn:
        c = conn.cursor()
        if department:
            rules = c.execute(
                "SELECT * FROM schedule_rules WHERE department=?", (department,)
            ).fetchall()
        else:
            rules = c.execute("SELECT * FROM schedule_rules").fetchall()

        rows = c.execute(
            """
        SELECT date, department, COALESCE(role,'') as role, COUNT(*) as scheduled
        FROM schedule_shifts
        WHERE date BETWEEN date(?) AND date(?) AND (code IS NULL OR code NOT IN ('休','例','OFF'))
        GROUP BY date, department, role
        """,
            (start, end),
        ).fetchall()
        scheduled_map = {
            (r["date"], r["department"], r["role"]): r["scheduled"] for r in rows
        }

        met = c.execute(
            """
        SELECT * FROM schedule_day_metrics WHERE date BETWEEN date(?) AND date(?)
        """,
            (start, end),
        ).fetchall()
        metric_map = {m["date"]: m for m in met}

    out: Dict[str, Dict[str, Dict[str, Any]]] = {}
    d = start_d
    while d <= end_d:
        dkey = d.isoformat()
        out.setdefault(dkey, {})
        for rule in rules:
            dep = rule["department"]
            role = rule["role"]
            out[dkey].setdefault(dep, {})
            r = occ_map.get(d, {})
            ratio = float(r.get("ratio") or 0.0)
            m = metric_map.get(dkey, {})
            guest = int(m.get("guest_count") or 0)
            bfast = int(m.get("breakfast_count") or 0)

            req = (
                (rule["base"] or 0)
                + (rule["per_occ"] or 0) * ratio
                + (rule["per_guest"] or 0) * guest
                + (rule["per_breakfast"] or 0) * bfast
            )
            req = math.ceil(req)
            req = max(req, int(rule["min_staff"] or 0))
            req = min(req, int(rule["max_staff"] or 9999))
            sch = scheduled_map.get((dkey, dep, role), 0)
            out[dkey][dep][role] = {
                "required": req,
                "scheduled": sch,
                "diff": sch - req,
                "ratio": ratio,
            }
        d += timedelta(days=1)
    return JSONResponse(out)


# === Excel 匯入（處理「所有工作表」；能辨識相同版型） ===
@router.post("/schedule/import/excel", include_in_schema=False)
async def import_schedule_excel(request: Request, file: UploadFile = File(...)):
    _require_perm(request, "schedule_manage")
    try:
        import openpyxl
    except Exception:
        return JSONResponse(
            {"ok": False, "error": "請先安裝 openpyxl：pip install openpyxl"},
            status_code=500,
        )

    from tempfile import NamedTemporaryFile

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        data = await file.read()
        tmp.write(data)
        tmp_path = tmp.name

    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    imported_sheets = 0
    with _db(PORTAL_DB) as conn:
        c = conn.cursor()

        def parse_ws(ws):
            # 版型假設：第3列是標頭（…/員編/英文/姓名），第4列=日期序、第5列=住房率、第6=預估房客、第7=預估早餐，第10列起=員工資料
            header_row, date_row, occ_row, guest_row, bfast_row, start_data_row = (
                3,
                4,
                5,
                6,
                7,
                10,
            )

            # 找「姓名」欄決定日期起始欄（姓名右邊一格開始）
            start_col = None
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=header_row, column=col).value
                if v and str(v).strip() in ("姓名", "name", "Name", "中文姓名"):
                    start_col = col + 1
                    break
            if not start_col:
                return False  # 不是這種版型

            # 讀日期欄
            days: List[Tuple[int, date]] = []
            for col in range(start_col, ws.max_column + 1):
                v = ws.cell(row=date_row, column=col).value
                if v is None:
                    continue
                try:
                    if isinstance(v, datetime):
                        d = v.date()
                    elif isinstance(v, (int, float)):
                        d = (datetime(1899, 12, 30) + timedelta(days=int(v))).date()
                    else:
                        d = datetime.fromisoformat(str(v)).date()
                    days.append((col, d))
                except:
                    continue
            if not days:  # 沒日期就略過
                return False

            # 寫入日指標
            for col, d in days:
                dkey = d.isoformat()
                occ = ws.cell(row=occ_row, column=col).value
                guest = ws.cell(row=guest_row, column=col).value
                bfast = ws.cell(row=bfast_row, column=col).value
                c.execute(
                    """
                INSERT INTO schedule_day_metrics(date, occ_ratio, guest_count, breakfast_count)
                VALUES(?,?,?,?)
                ON CONFLICT(date) DO UPDATE SET occ_ratio=excluded.occ_ratio, guest_count=excluded.guest_count, breakfast_count=excluded.breakfast_count
                """,
                    (
                        dkey,
                        float(occ) if occ not in (None, "") else None,
                        int(guest or 0),
                        int(bfast or 0),
                    ),
                )

            # 寫入員工班次
            for r in range(start_data_row, ws.max_row + 1):
                dept = ws.cell(row=r, column=1).value
                dept1 = ws.cell(row=r, column=2).value
                unit = ws.cell(row=r, column=3).value
                title = ws.cell(row=r, column=4).value
                empno = ws.cell(row=r, column=5).value
                en = ws.cell(row=r, column=6).value
                name = ws.cell(row=r, column=7).value
                if not any([dept, dept1, name, empno]):  # 空列
                    continue
                for col, d in days:
                    code = ws.cell(row=r, column=col).value
                    if code in (None, ""):
                        continue
                    code_str = str(code).replace(".", "").strip()
                    start_t, end_t = None, None
                    if re.match(r"^\d{1,2}[-~]\d{1,2}$", code_str):
                        hh1, hh2 = re.split(r"[-~]", code_str)
                        start_t = f"{int(hh1):02d}:00"
                        end_t = f"{int(hh2):02d}:00"
                    elif code_str in ("休", "OFF", "例"):
                        start_t = end_t = None
                    c.execute(
                        """
                    INSERT INTO schedule_shifts(emp_id, emp_name, department, department_1, role, date, code, start_time, end_time, notes, created_by)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)
                    """,
                        (
                            str(empno or ""),
                            str(name or en or ""),
                            str(dept or ""),
                            str(dept1 or ""),
                            str(title or ""),
                            d.isoformat(),
                            code_str,
                            start_t,
                            end_t,
                            unit or None,
                            request.cookies.get("user") or "import",
                        ),
                    )
            return True

        for name in wb.sheetnames:
            ws = wb[name]
            if parse_ws(ws):
                imported_sheets += 1

        conn.commit()

    if imported_sheets == 0:
        return JSONResponse(
            {
                "ok": False,
                "error": "沒有可辨識的工作表版型（需符合：第3列標頭、第4列日期、第5列住房率、第10列起員工）",
            },
            status_code=400,
        )
    return JSONResponse({"ok": True, "sheets": imported_sheets})


# === Page ===
@router.get("/schedule", response_class=HTMLResponse, include_in_schema=False)
async def schedule_page(
    request: Request,
    y: Optional[int] = None,
    m: Optional[int] = None,
    department: Optional[str] = None,
    department_1: Optional[str] = None,
):
    _require_perm(request, "schedule_view")
    today = date.today()
    y = y or today.year
    m = m or today.month
    first = date(y, m, 1)

    # 週一開頭的 6 週格
    start_grid = first - timedelta(days=(first.weekday() - 0) % 7)
    weeks: List[List[date]] = []
    d = start_grid
    for _ in range(6):
        row: List[date] = []
        for __ in range(7):
            row.append(d)
            d += timedelta(days=1)
        weeks.append(row)
    end_grid = weeks[-1][-1]

    # 上/下月
    prev_month = (first.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_month = (first.replace(day=28) + timedelta(days=4)).replace(day=1)

    # 住房率
    occ_by_day = _load_month_occupancy(start_grid, end_grid)

    # 部門清單（即使沒員工也能選）
    dept_opts, dept1_opts = _load_dept_lists()
    # 員工（篩部門/組別）
    employees = _load_employees(department, department_1)

    user = request.cookies.get("user")
    role = request.cookies.get("role")
    permissions = request.cookies.get("permissions")
    ctx = get_base_context(request, user, role, permissions)
    ctx.update(
        {
            "page_title": "員工排班表",
            "year": y,
            "month": m,
            "weeks": weeks,
            "grid_start": start_grid,
            "grid_end": end_grid,
            "prev_y": prev_month.year,
            "prev_m": prev_month.month,
            "next_y": next_month.year,
            "next_m": next_month.month,
            "occ_by_day": occ_by_day,
            "employees": employees,
            "dept_opts": dept_opts,
            "dept1_opts": dept1_opts,
            "active_menu": "schedule",
            "can_manage": _has_perm(request, "schedule_manage"),
            "can_numeric": _has_perm(request, "schedule_numeric"),
            "filter_department": department or "",
            "filter_department_1": department_1 or "",
        }
    )
    return templates.TemplateResponse("schedule/index.html", ctx)
